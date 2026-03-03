"""
월별 매출 보고서 자동 생성 스크립트
openpyxl 핵심 기능 데모:
  - 셀 값 입력 (문자열, 숫자, 수식)
  - 스타일링 (Font, PatternFill, Alignment, Border)
  - 값 기반 조건부 색상
  - BarChart 자동 삽입
  - 열 너비 자동 조정
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
import os

# ── 샘플 데이터 ──────────────────────────────────────────
MONTHLY_DATA = [
    ("1월",  "서울",  4_200_000, 3_800_000, 4_100_000),
    ("2월",  "서울",  3_900_000, 4_100_000, 3_700_000),
    ("3월",  "서울",  5_300_000, 4_800_000, 5_100_000),
    ("4월",  "서울",  4_700_000, 5_200_000, 4_900_000),
    ("5월",  "서울",  6_100_000, 5_700_000, 6_300_000),
    ("6월",  "서울",  5_800_000, 6_000_000, 5_600_000),
    ("7월",  "부산",  3_400_000, 3_200_000, 3_600_000),
    ("8월",  "부산",  3_700_000, 3_900_000, 3_500_000),
    ("9월",  "부산",  4_800_000, 4_500_000, 4_700_000),
    ("10월", "부산",  5_200_000, 5_000_000, 5_400_000),
    ("11월", "부산",  6_500_000, 6_200_000, 6_800_000),
    ("12월", "부산",  7_100_000, 7_300_000, 6_900_000),
]
HEADERS = ["월", "지점", "온라인 매출", "오프라인 매출", "목표 매출", "합계", "달성률"]


# ── 스타일 정의 ──────────────────────────────────────────
def header_style():
    """헤더 행 스타일"""
    return {
        "font": Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", fgColor="2E4057"),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": thin_border(),
    }


def data_style(is_even: bool):
    """데이터 행 배경 (줄무늬)"""
    fill_color = "EFF3F8" if is_even else "FFFFFF"
    return {
        "font": Font(name="맑은 고딕", size=10),
        "fill": PatternFill("solid", fgColor=fill_color),
        "alignment": Alignment(horizontal="center", vertical="center"),
        "border": thin_border(),
    }


def number_style(is_even: bool):
    """숫자 셀 스타일 (오른쪽 정렬 + 천 단위)"""
    fill_color = "EFF3F8" if is_even else "FFFFFF"
    return {
        "font": Font(name="맑은 고딕", size=10),
        "fill": PatternFill("solid", fgColor=fill_color),
        "alignment": Alignment(horizontal="right", vertical="center"),
        "border": thin_border(),
        "number_format": "#,##0",
    }


def thin_border():
    side = Side(style="thin", color="C8D0D8")
    return Border(left=side, right=side, top=side, bottom=side)


def achievement_fill(rate: float) -> PatternFill:
    """달성률에 따른 배경색"""
    if rate >= 1.0:
        return PatternFill("solid", fgColor="C8EFCB")   # 초록
    elif rate >= 0.9:
        return PatternFill("solid", fgColor="FFF2CC")   # 노랑
    else:
        return PatternFill("solid", fgColor="FCE4D6")   # 빨강


def achievement_font(rate: float) -> Font:
    if rate >= 1.0:
        color = "1A7A1A"
    elif rate >= 0.9:
        color = "7A5A00"
    else:
        color = "A0230A"
    return Font(name="맑은 고딕", bold=True, size=10, color=color)


# ── 셀에 스타일 일괄 적용 ────────────────────────────────
def apply_style(cell, styles: dict):
    for attr, val in styles.items():
        setattr(cell, attr, val)


# ── 보고서 생성 ──────────────────────────────────────────
def create_report(output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "2024 매출 보고서"

    # 1. 제목 행 (병합)
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "2024년 월별 매출 보고서"
    title_cell.font = Font(name="맑은 고딕", bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="1A2E44")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # 2. 헤더 행
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        apply_style(cell, header_style())
    ws.row_dimensions[2].height = 28

    # 3. 데이터 행
    for row_idx, (month, branch, online, offline, target) in enumerate(MONTHLY_DATA, start=3):
        is_even = (row_idx % 2 == 0)
        total_formula = f"=C{row_idx}+D{row_idx}"
        rate_formula = f"=F{row_idx}/E{row_idx}"

        row_data = [month, branch, online, offline, target, total_formula, rate_formula]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_idx <= 2:
                apply_style(cell, data_style(is_even))
            elif col_idx <= 6:
                apply_style(cell, number_style(is_even))
            else:
                # 달성률 열: 수식 결과로 조건부 색상 결정
                rate = (online + offline) / target
                cell.font = achievement_font(rate)
                cell.fill = achievement_fill(rate)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border()
                cell.number_format = "0.0%"

        ws.row_dimensions[row_idx].height = 22

    # 4. 합계/평균 행
    summary_row = len(MONTHLY_DATA) + 3
    ws.merge_cells(f"A{summary_row}:B{summary_row}")
    label_cell = ws[f"A{summary_row}"]
    label_cell.value = "합계 / 평균"
    label_cell.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    label_cell.fill = PatternFill("solid", fgColor="2E4057")
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = thin_border()
    ws[f"B{summary_row}"].border = thin_border()

    data_start, data_end = 3, 3 + len(MONTHLY_DATA) - 1
    for col_idx in range(3, 7):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=summary_row, column=col_idx)
        cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{data_end})"
        cell.font = Font(name="맑은 고딕", bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="DAE3F0")
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = thin_border()
        cell.number_format = "#,##0"

    avg_cell = ws.cell(row=summary_row, column=7)
    avg_cell.value = f"=AVERAGE(G3:G{data_end})"
    avg_cell.font = Font(name="맑은 고딕", bold=True, size=10)
    avg_cell.fill = PatternFill("solid", fgColor="DAE3F0")
    avg_cell.alignment = Alignment(horizontal="center", vertical="center")
    avg_cell.border = thin_border()
    avg_cell.number_format = "0.0%"
    ws.row_dimensions[summary_row].height = 24

    # 5. 열 너비 조정
    col_widths = [8, 8, 14, 14, 14, 14, 10]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 6. 차트 생성 (온라인 + 오프라인 월별 막대 차트)
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "월별 온라인/오프라인 매출"
    chart.y_axis.title = "매출 (원)"
    chart.x_axis.title = "월"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    online_ref = Reference(ws, min_col=3, min_row=2, max_row=data_end)
    offline_ref = Reference(ws, min_col=4, min_row=2, max_row=data_end)
    month_ref = Reference(ws, min_col=1, min_row=3, max_row=data_end)

    chart.add_data(online_ref, titles_from_data=True)
    chart.add_data(offline_ref, titles_from_data=True)
    chart.set_categories(month_ref)

    chart.series[0].graphicalProperties.solidFill = "4472C4"
    chart.series[1].graphicalProperties.solidFill = "ED7D31"

    ws.add_chart(chart, "I2")

    # 7. 시트 기본 설정
    ws.freeze_panes = "A3"          # 헤더 고정
    ws.sheet_view.showGridLines = False

    # 8. 저장
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"✓ 보고서 저장 완료: {output_path}")
    print(f"  - 데이터 행: {len(MONTHLY_DATA)}개")
    print(f"  - 시트: {ws.title}")
    print(f"  - 차트: 월별 온라인/오프라인 막대 차트")


if __name__ == "__main__":
    output_path = os.path.join("output", "monthly_sales_report.xlsx")
    create_report(output_path)
